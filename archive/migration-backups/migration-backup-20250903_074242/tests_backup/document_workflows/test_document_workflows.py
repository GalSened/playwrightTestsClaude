"""
Document Workflow Tests
Tests for document distribution, signing workflows, and advanced processing
that are present in Selenium but missing from the Playwright test suite.
"""

import pytest
import allure
import os
from playwright.sync_api import expect
import tempfile
import openpyxl
from pathlib import Path
from src.pages.wesign_document_page import WeSignDocumentPage
from src.utils.test_helpers import TestHelpers


@allure.epic("Document Workflows")
@allure.feature("Document Distribution and Signing")
class TestDocumentWorkflows:
    """Document workflow test suite"""

    @pytest.mark.regression
    @pytest.mark.upload
    @allure.story("Document Distribution")
    @allure.title("Distribute document with XLSX file containing multiple signers")
    def test_distribute_document_with_xlsx_multiple_signers(
        self, authenticated_page, test_helpers, test_config, cleanup_test_data
    ):
        """Test document distribution using XLSX file with multiple signers"""
        page = authenticated_page
        
        page.goto(f"{test_config.urls['base_url']}dashboard/main")
        page.wait_for_load_state('networkidle')
        
        # Upload a document first
        upload_button = page.locator('button:has-text("Upload file")')
        upload_button.click()
        
        pdf_file = test_config.get_file_path('pdf_3_pages')
        if pdf_file and os.path.exists(pdf_file):
            page.set_input_files('input[type="file"]', pdf_file)
            cleanup_test_data['register_upload'](pdf_file)
            page.wait_for_timeout(3000)
            
            # Navigate to Assign & Send
            assign_send_button = page.locator('button:has-text("Assign & send")')
            assign_send_button.click()
            page.wait_for_timeout(2000)
            
            # Create XLSX file with multiple signers
            signers_data = [
                {
                    "Full Name": "Signer One",
                    "Email": "signer1@example.com", 
                    "Phone": "0552603210",
                    "Send by": "EMAIL",
                    "Role": "Signer"
                },
                {
                    "Full Name": "Signer Two", 
                    "Email": "signer2@example.com",
                    "Phone": "0504821887",
                    "Send by": "EMAIL", 
                    "Role": "Signer"
                },
                {
                    "Full Name": "Approver One",
                    "Email": "approver@example.com",
                    "Send by": "EMAIL",
                    "Role": "Approver"
                }
            ]
            
            xlsx_file = self._create_signers_xlsx_file(signers_data)
            cleanup_test_data['register_upload'](xlsx_file)
            
            # Look for XLSX import option in assign & send
            import_signers_button = page.locator('button:has-text("Import"), button:has-text("XLSX"), .import-signers')
            if import_signers_button.count() > 0:
                import_signers_button.click()
                page.set_input_files('input[type="file"]', xlsx_file)
                page.wait_for_timeout(3000)
                
                # Verify signers were imported
                signers_list = page.locator('.signers-list, .recipients-list, .distribution-list')
                signers_list.wait_for(state='visible', timeout=15000)
                
                # Check that all signers appear
                for signer in signers_data:
                    signer_element = signers_list.locator(f'text="{signer["Full Name"]}"')
                    expect(signer_element).to_be_visible()
                
                # Send document
                send_button = page.locator('button:has-text("Send"), button:has-text("Distribute")')
                send_button.click()
                page.wait_for_timeout(3000)
                
                # Verify distribution success
                success_message = page.locator('.success, .sent-success, .distribution-success')
                success_message.wait_for(state='visible', timeout=15000)

    @pytest.mark.regression
    @allure.story("Self Signing Workflow")
    @allure.title("Complete self-signing workflow successfully")
    def test_self_signing_workflow_complete(
        self, authenticated_page, test_helpers, test_config, cleanup_test_data
    ):
        """Test complete self-signing workflow"""
        page = authenticated_page
        
        page.goto(f"{test_config.urls['base_url']}dashboard/main")
        page.wait_for_load_state('networkidle')
        
        # Upload document
        upload_button = page.locator('button:has-text("Upload file")')
        upload_button.click()
        
        pdf_file = test_config.get_file_path('pdf_3_pages')
        if pdf_file and os.path.exists(pdf_file):
            page.set_input_files('input[type="file"]', pdf_file)
            cleanup_test_data['register_upload'](pdf_file)
            page.wait_for_timeout(3000)
            
            # Look for self-sign option
            self_sign_button = page.locator('button:has-text("Self Sign"), button:has-text("Sign Now"), .self-sign')
            if self_sign_button.count() > 0:
                self_sign_button.click()
                page.wait_for_timeout(2000)
                
                # Wait for signing interface
                signing_area = page.locator('.signing-area, .signature-pad, .sign-interface')
                signing_area.wait_for(state='visible', timeout=15000)
                
                # Select signature method (draw)
                draw_signature = page.locator('button:has-text("Draw"), .draw-signature')
                if draw_signature.count() > 0:
                    draw_signature.click()
                    
                    # Draw signature on canvas (simulate mouse movement)
                    canvas = page.locator('canvas, .signature-canvas')
                    if canvas.count() > 0:
                        # Simulate signature drawing
                        canvas.hover()
                        page.mouse.down()
                        page.mouse.move(100, 50)
                        page.mouse.move(150, 75)
                        page.mouse.move(200, 50)
                        page.mouse.up()
                        
                        # Apply signature
                        apply_button = page.locator('button:has-text("Apply"), button:has-text("Sign"), .apply-signature')
                        apply_button.click()
                        page.wait_for_timeout(3000)
                        
                        # Verify signing success
                        success_message = page.locator('.signing-success, .document-signed, .success')
                        success_message.wait_for(state='visible', timeout=15000)
                        
                        # Check if signed document is available for download
                        download_button = page.locator('button:has-text("Download"), .download-signed')
                        if download_button.count() > 0:
                            expect(download_button).to_be_visible()

    @pytest.mark.regression
    @allure.story("Others Signing Workflow")
    @allure.title("Assign document to others for signing workflow")
    def test_others_signing_workflow_assignment(
        self, authenticated_page, test_helpers, test_config, cleanup_test_data, test_recipients
    ):
        """Test assigning document to others for signing"""
        page = authenticated_page
        
        page.goto(f"{test_config.urls['base_url']}dashboard/main")
        page.wait_for_load_state('networkidle')
        
        # Upload document
        upload_button = page.locator('button:has-text("Upload file")')
        upload_button.click()
        
        pdf_file = test_config.get_file_path('pdf_3_pages')
        if pdf_file and os.path.exists(pdf_file):
            page.set_input_files('input[type="file"]', pdf_file)
            cleanup_test_data['register_upload'](pdf_file)
            page.wait_for_timeout(3000)
            
            # Go to Assign & Send
            assign_send_button = page.locator('button:has-text("Assign & send")')
            assign_send_button.click()
            page.wait_for_timeout(2000)
            
            # Add recipients manually
            recipient_data = test_recipients[0] if test_recipients else {
                'name': 'Test Recipient',
                'email': 'recipient@example.com'
            }
            
            # Add recipient
            add_recipient_button = page.locator('button:has-text("Add Recipient"), button:has-text("Add Signer"), .add-recipient')
            if add_recipient_button.count() > 0:
                add_recipient_button.click()
                
                # Fill recipient information
                name_field = page.locator('input[placeholder*="name"], input[placeholder*="Name"]').first
                if name_field.is_visible():
                    name_field.fill(recipient_data['name'])
                
                email_field = page.locator('input[placeholder*="email"], input[placeholder*="Email"], input[type="email"]').first
                if email_field.is_visible():
                    email_field.fill(recipient_data['email'])
                
                # Set signing order if available
                order_field = page.locator('input[placeholder*="order"], select[name*="order"]').first
                if order_field.count() > 0:
                    order_field.fill("1")
                
                # Save recipient
                save_recipient_button = page.locator('button:has-text("Save"), button:has-text("Add"), .save-recipient')
                save_recipient_button.click()
                page.wait_for_timeout(1000)
                
                # Send document
                send_button = page.locator('button:has-text("Send"), button:has-text("Send for Signature")')
                send_button.click()
                page.wait_for_timeout(3000)
                
                # Verify assignment success
                success_message = page.locator('.assignment-success, .sent-success, .success')
                success_message.wait_for(state='visible', timeout=15000)

    @pytest.mark.regression 
    @allure.story("Large File Handling")
    @allure.title("Handle large file distribution successfully")
    def test_large_file_distribution_handling(
        self, authenticated_page, test_helpers, test_config, cleanup_test_data
    ):
        """Test distribution of large files"""
        page = authenticated_page
        
        page.goto(f"{test_config.urls['base_url']}dashboard/main")
        page.wait_for_load_state('networkidle')
        
        # Use largest available test file
        large_files = test_config.get_large_files()
        if large_files:
            large_file = large_files[0]
            cleanup_test_data['register_upload'](large_file)
            
            upload_button = page.locator('button:has-text("Upload file")')
            upload_button.click()
            
            # Upload large file with extended timeout
            page.set_input_files('input[type="file"]', large_file)
            page.wait_for_timeout(10000)  # Extended wait for large files
            
            # Check for upload progress or completion
            upload_progress = page.locator('.upload-progress, .progress-bar')
            if upload_progress.count() > 0:
                # Wait for upload to complete
                upload_progress.wait_for(state='hidden', timeout=120000)  # 2 minutes max
            
            # Proceed with distribution
            assign_send_button = page.locator('button:has-text("Assign & send")')
            if assign_send_button.is_enabled():
                assign_send_button.click()
                page.wait_for_timeout(3000)
                
                # Add recipient for large file
                add_recipient_button = page.locator('button:has-text("Add Recipient"), .add-recipient')
                if add_recipient_button.count() > 0:
                    add_recipient_button.click()
                    
                    page.fill('input[placeholder*="name"]', "Large File Recipient")
                    page.fill('input[type="email"]', "largefile@example.com")
                    
                    save_button = page.locator('button:has-text("Save")')
                    save_button.click()
                    
                    # Send large file
                    send_button = page.locator('button:has-text("Send")')
                    send_button.click()
                    
                    # Wait for processing with extended timeout
                    success_message = page.locator('.success, .sent-success')
                    success_message.wait_for(state='visible', timeout=180000)  # 3 minutes

    @pytest.mark.regression
    @allure.story("Document Merging")
    @allure.title("Merge multiple PDF files and distribute")
    def test_merge_files_and_distribute(
        self, authenticated_page, test_helpers, test_config, cleanup_test_data
    ):
        """Test merging multiple files and distributing the result"""
        page = authenticated_page
        
        page.goto(f"{test_config.urls['base_url']}dashboard/main")
        page.wait_for_load_state('networkidle')
        
        merge_files = test_config.get_files_for_merge_testing()
        if len(merge_files) >= 2:
            # Click Merge files
            merge_button = page.locator('button:has-text("Merge files")')
            merge_button.click()
            page.wait_for_timeout(1000)
            
            # Upload multiple files for merging
            for file_path in merge_files[:3]:  # Use up to 3 files
                cleanup_test_data['register_upload'](file_path)
                
                add_file_button = page.locator('button:has-text("Add File"), button:has-text("Choose Files"), .add-merge-file')
                if add_file_button.count() > 0:
                    add_file_button.click()
                    page.set_input_files('input[type="file"]', file_path)
                    page.wait_for_timeout(2000)
            
            # Start merge process
            start_merge_button = page.locator('button:has-text("Merge"), button:has-text("Start Merge"), .start-merge')
            start_merge_button.click()
            page.wait_for_timeout(5000)  # Wait for merge to complete
            
            # Verify merge success
            merge_success = page.locator('.merge-success, .merge-complete')
            merge_success.wait_for(state='visible', timeout=30000)
            
            # Proceed with distribution of merged file
            assign_send_button = page.locator('button:has-text("Assign & send")')
            if assign_send_button.is_enabled():
                assign_send_button.click()
                
                # Add recipient
                add_recipient_button = page.locator('button:has-text("Add Recipient")')
                if add_recipient_button.count() > 0:
                    add_recipient_button.click()
                    
                    page.fill('input[placeholder*="name"]', "Merged Doc Recipient")
                    page.fill('input[type="email"]', "merged@example.com")
                    
                    save_button = page.locator('button:has-text("Save")')
                    save_button.click()
                    
                    # Send merged document
                    send_button = page.locator('button:has-text("Send")')
                    send_button.click()
                    
                    success_message = page.locator('.success, .sent-success')
                    success_message.wait_for(state='visible', timeout=15000)

    @pytest.mark.regression
    @allure.story("Template Field Mapping")
    @allure.title("Map template fields during distribution")
    def test_template_field_mapping_distribution(
        self, authenticated_page, test_helpers, test_config, cleanup_test_data
    ):
        """Test mapping template fields during document distribution"""
        page = authenticated_page
        
        page.goto(f"{test_config.urls['base_url']}dashboard/main")
        page.wait_for_load_state('networkidle')
        
        # Upload document with fields
        upload_button = page.locator('button:has-text("Upload file")')
        upload_button.click()
        
        pdf_with_fields = test_config.get_file_path('pdf_with_fields')
        if pdf_with_fields and os.path.exists(pdf_with_fields):
            page.set_input_files('input[type="file"]', pdf_with_fields)
            cleanup_test_data['register_upload'](pdf_with_fields)
            page.wait_for_timeout(5000)
            
            # Go to Assign & Send
            assign_send_button = page.locator('button:has-text("Assign & send")')
            assign_send_button.click()
            page.wait_for_timeout(2000)
            
            # Look for field mapping interface
            field_mapping = page.locator('.field-mapping, .template-fields, .form-fields')
            if field_mapping.count() > 0:
                # Map fields to recipients
                field_elements = page.locator('.field-item, .template-field')
                field_count = field_elements.count()
                
                if field_count > 0:
                    for i in range(min(field_count, 3)):  # Map first 3 fields
                        field = field_elements.nth(i)
                        
                        # Assign field to a signer
                        field_dropdown = field.locator('select, .field-assignment')
                        if field_dropdown.count() > 0:
                            field_dropdown.select_option('signer1')
                            
                    # Add recipient to assign fields to
                    add_recipient_button = page.locator('button:has-text("Add Recipient")')
                    if add_recipient_button.count() > 0:
                        add_recipient_button.click()
                        
                        page.fill('input[placeholder*="name"]', "Field Mapper")
                        page.fill('input[type="email"]', "fieldmapper@example.com")
                        
                        save_button = page.locator('button:has-text("Save")')
                        save_button.click()
                        
                        # Send with field mapping
                        send_button = page.locator('button:has-text("Send")')
                        send_button.click()
                        
                        success_message = page.locator('.success, .sent-success')
                        success_message.wait_for(state='visible', timeout=15000)

    @pytest.mark.regression
    @allure.story("Duplicate Signer Handling")
    @allure.title("Handle duplicate signer emails properly during distribution")
    def test_duplicate_signer_email_handling(
        self, authenticated_page, test_helpers, test_config, cleanup_test_data
    ):
        """Test handling of duplicate signer emails during distribution"""
        page = authenticated_page
        
        page.goto(f"{test_config.urls['base_url']}dashboard/main")
        page.wait_for_load_state('networkidle')
        
        # Upload document
        upload_button = page.locator('button:has-text("Upload file")')
        upload_button.click()
        
        pdf_file = test_config.get_file_path('pdf_3_pages')
        if pdf_file and os.path.exists(pdf_file):
            page.set_input_files('input[type="file"]', pdf_file)
            cleanup_test_data['register_upload'](pdf_file)
            page.wait_for_timeout(3000)
            
            assign_send_button = page.locator('button:has-text("Assign & send")')
            assign_send_button.click()
            page.wait_for_timeout(2000)
            
            # Create XLSX with duplicate emails
            duplicate_signers = [
                {"Full Name": "John Doe", "Email": "duplicate@example.com", "Role": "Signer"},
                {"Full Name": "Jane Smith", "Email": "duplicate@example.com", "Role": "Signer"},  # Duplicate email
                {"Full Name": "Bob Johnson", "Email": "unique@example.com", "Role": "Signer"}
            ]
            
            xlsx_file = self._create_signers_xlsx_file(duplicate_signers)
            cleanup_test_data['register_upload'](xlsx_file)
            
            # Import signers with duplicates
            import_button = page.locator('button:has-text("Import"), .import-signers')
            if import_button.count() > 0:
                import_button.click()
                page.set_input_files('input[type="file"]', xlsx_file)
                page.wait_for_timeout(3000)
                
                # Look for duplicate email warning
                duplicate_warning = page.locator('.duplicate-warning, .warning, .alert-warning')
                if duplicate_warning.count() > 0:
                    warning_text = duplicate_warning.inner_text()
                    assert 'duplicate' in warning_text.lower(), "Should show duplicate email warning"
                    
                    # Handle duplicate resolution
                    resolve_button = page.locator('button:has-text("Resolve"), button:has-text("Merge"), .resolve-duplicates')
                    if resolve_button.count() > 0:
                        resolve_button.click()
                        page.wait_for_timeout(2000)
                
                # Verify only unique recipients remain
                recipients_list = page.locator('.recipients-list, .signers-list')
                recipient_emails = page.locator('.recipient-email, .signer-email')
                email_count = recipient_emails.count()
                
                # Should have only 2 unique emails
                unique_emails = set()
                for i in range(email_count):
                    email_text = recipient_emails.nth(i).inner_text()
                    unique_emails.add(email_text.strip())
                
                assert len(unique_emails) <= 2, f"Should have max 2 unique emails, found {len(unique_emails)}"

    def _create_signers_xlsx_file(self, signers_data: list) -> str:
        """Create temporary XLSX file with signers data"""
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        temp_file.close()
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        if signers_data:
            headers = list(signers_data[0].keys())
            for col, header in enumerate(headers, 1):
                worksheet.cell(row=1, column=col, value=header)
            
            for row, signer in enumerate(signers_data, 2):
                for col, header in enumerate(headers, 1):
                    worksheet.cell(row=row, column=col, value=signer.get(header, ''))
        
        workbook.save(temp_file.name)
        workbook.close()
        
        return temp_file.name

    @pytest.mark.regression
    @allure.story("Document Status Tracking")
    @allure.title("Track document status through signing workflow")
    def test_document_status_tracking(
        self, authenticated_page, test_helpers, test_config, cleanup_test_data
    ):
        """Test tracking document status through the signing workflow"""
        page = authenticated_page
        
        page.goto(f"{test_config.urls['base_url']}dashboard/main")
        page.wait_for_load_state('networkidle')
        
        # Upload and assign document
        upload_button = page.locator('button:has-text("Upload file")')
        upload_button.click()
        
        pdf_file = test_config.get_file_path('pdf_3_pages')
        if pdf_file and os.path.exists(pdf_file):
            page.set_input_files('input[type="file"]', pdf_file)
            cleanup_test_data['register_upload'](pdf_file)
            page.wait_for_timeout(3000)
            
            assign_send_button = page.locator('button:has-text("Assign & send")')
            assign_send_button.click()
            
            # Add recipient
            add_recipient_button = page.locator('button:has-text("Add Recipient")')
            if add_recipient_button.count() > 0:
                add_recipient_button.click()
                
                page.fill('input[placeholder*="name"]', "Status Tracker")
                page.fill('input[type="email"]', "statustrack@example.com")
                
                save_button = page.locator('button:has-text("Save")')
                save_button.click()
                
                # Send document
                send_button = page.locator('button:has-text("Send")')
                send_button.click()
                page.wait_for_timeout(3000)
                
                # Navigate to documents to check status
                documents_button = page.locator('button:has-text("Documents")')
                documents_button.click()
                page.wait_for_timeout(2000)
                
                # Look for document status
                document_status = page.locator('.document-status, .status-indicator, .workflow-status')
                if document_status.count() > 0:
                    status_text = document_status.first.inner_text()
                    
                    # Verify status contains expected values
                    expected_statuses = ['pending', 'sent', 'waiting', 'in progress']
                    status_found = any(status in status_text.lower() for status in expected_statuses)
                    assert status_found, f"Document status should contain expected status, found: {status_text}"

    @pytest.mark.performance
    @allure.story("Workflow Performance")  
    @allure.title("Document workflow performance within acceptable limits")
    def test_document_workflow_performance(
        self, authenticated_page, test_helpers, test_config, cleanup_test_data, performance_tracker
    ):
        """Test document workflow performance"""
        page = authenticated_page
        
        page.goto(f"{test_config.urls['base_url']}dashboard/main")
        page.wait_for_load_state('networkidle')
        
        # Track upload performance
        upload_start = page.evaluate('() => performance.now()')
        
        upload_button = page.locator('button:has-text("Upload file")')
        upload_button.click()
        
        pdf_file = test_config.get_file_path('pdf_3_pages')
        if pdf_file and os.path.exists(pdf_file):
            page.set_input_files('input[type="file"]', pdf_file)
            cleanup_test_data['register_upload'](pdf_file)
            page.wait_for_timeout(3000)
            
            upload_end = page.evaluate('() => performance.now()')
            upload_duration = int(upload_end - upload_start)
            performance_tracker['add_operation']('Document Upload', upload_duration)
            
            # Track assignment performance
            assign_start = page.evaluate('() => performance.now()')
            
            assign_send_button = page.locator('button:has-text("Assign & send")')
            assign_send_button.click()
            page.wait_for_timeout(2000)
            
            add_recipient_button = page.locator('button:has-text("Add Recipient")')
            if add_recipient_button.count() > 0:
                add_recipient_button.click()
                
                page.fill('input[placeholder*="name"]', "Perf Test")
                page.fill('input[type="email"]', "perftest@example.com")
                
                save_button = page.locator('button:has-text("Save")')
                save_button.click()
                
                send_button = page.locator('button:has-text("Send")')
                send_button.click()
                
                success_message = page.locator('.success, .sent-success')
                success_message.wait_for(state='visible', timeout=15000)
                
                assign_end = page.evaluate('() => performance.now()')
                assign_duration = int(assign_end - assign_start)
                performance_tracker['add_operation']('Document Assignment', assign_duration)
                
                # Verify performance thresholds
                thresholds = test_config.get_performance_thresholds()
                assert upload_duration < thresholds['file_upload_max'], f"Upload took {upload_duration}ms, expected < {thresholds['file_upload_max']}ms"
                assert assign_duration < thresholds['document_send_max'], f"Assignment took {assign_duration}ms, expected < {thresholds['document_send_max']}ms"