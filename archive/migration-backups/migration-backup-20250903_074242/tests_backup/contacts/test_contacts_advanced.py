"""
Advanced Contact Management Tests
Tests for XLSX import, seal functionality, tag management, and validation scenarios
that are present in Selenium but missing from the basic Playwright contact tests.
"""

import pytest
import allure
import os
from pathlib import Path
import openpyxl
import tempfile
from playwright.sync_api import expect
from src.pages.wesign_document_page import WeSignDocumentPage
from src.utils.test_helpers import TestHelpers
from src.utils.locators import WeSignLocators, LocatorHelper


@allure.epic("Contact Management")
@allure.feature("Advanced Contact Operations")
class TestAdvancedContacts:
    """Advanced contact management test suite"""
    @pytest.mark.upload
    @pytest.mark.regression
    @allure.story("XLSX Contact Import")
    @allure.title("Import valid XLSX file with email and SMS contacts successfully")
    def test_contacts_import_valid_xlsx_file_with_email_and_sms_success_english(
        self, authenticated_page, test_helpers, test_config, cleanup_test_data
    ):
        """Test importing XLSX file with both email and SMS contacts in English interface"""
        page = authenticated_page
        
        # Navigate to contacts page (try multiple navigation methods)
        page.wait_for_load_state('networkidle')
        
        # Try direct URL first
        contacts_url = f"{test_config.urls['base_url']}dashboard/contacts"
        page.goto(contacts_url)
        page.wait_for_load_state('networkidle')
        page.wait_for_timeout(2000)
        
        # Check if we need to navigate via menu instead
        if 'main' in page.url or not 'contacts' in page.url:
            # Use centralized locator helper for navigation
            locator_helper = LocatorHelper(page)
            
            if not locator_helper.click_with_fallbacks('navigation', 'contacts_nav', timeout=10000):
                allure.attach(f"Failed to navigate to contacts page. Current URL: {page.url}", 
                             name="Navigation Error", attachment_type=allure.attachment_type.TEXT)
        
        # Final validation - check for contacts page indicators
        page_indicators = [
            '[data-page="contacts"]',
            '.contacts-page',
            'h1:has-text("Contacts")',
            'h1:has-text("אנשי קשר")',  # Hebrew
            '.page-header:has-text("Contacts")',
            '[data-testid="contacts-page"]',
            # More generic indicators
            'button:has-text("Import")',
            'input[type="file"]',
            '.contact-list',
            '.contacts-table'
        ]
        
        page_found = False
        for indicator in page_indicators:
            try:
                page.locator(indicator).wait_for(state='visible', timeout=3000)
                page_found = True
                break
            except Exception:
                continue
        
        if not page_found:
            page.screenshot(path='artifacts/screenshots/navigation_debug.png')
            # Instead of failing, continue and let the import button check handle it
            allure.attach(f"Warning: Contacts page indicators not found. Current URL: {page.url}", name="Navigation Warning", attachment_type=allure.attachment_type.TEXT)
        
        # Create temporary XLSX file with test data
        test_xlsx = self._create_test_xlsx_file([
            {"Full Name": "Test User 1", "Email": "test1@example.com", "Phone": "0552603210", "Send by": "EMAIL"},
            {"Full Name": "Test User 2", "Phone": "0504821887", "Send by": "SMS"},
            {"Full Name": "Test User 3", "Email": "test3@example.com", "Phone": "0551234567", "Send by": "EMAIL"}
        ])
        cleanup_test_data['register_upload'](test_xlsx)
        
        # Click Import Excel button using centralized locators
        locator_helper = LocatorHelper(page)
        
        try:
            import_button = locator_helper.wait_for_element_with_fallbacks('contacts', 'import_excel_button', timeout=10000)
            import_button.click()
        except TimeoutError:
            # Take screenshot for debugging
            page.screenshot(path='artifacts/screenshots/contacts_page_debug.png')
            pytest.fail(f"Import Excel button not found on contacts page. Current URL: {page.url}")
        
        # Upload file
        page.set_input_files('input[type="file"]', test_xlsx)
        page.wait_for_timeout(2000)  # Wait for upload processing
        
        # Verify upload success by checking for imported contacts in table
        # Wait for table to load/update with new contacts
        page.wait_for_timeout(3000)
        
        # First try to find success message, but don't fail if not found
        try:
            success_element = locator_helper.wait_for_element_with_fallbacks('status', 'success_message', timeout=5000)
            expect(success_element).to_be_visible()
            allure.attach("Success message found and visible", name="Success Message Verification", attachment_type=allure.attachment_type.TEXT)
        except TimeoutError:
            # Success message not found, but continue with verifying the actual import result
            allure.attach("Success message not visible, but will verify import through table data", name="Success Verification Method", attachment_type=allure.attachment_type.TEXT)
        
        # Verify contacts appear in table with strong assertions
        page.wait_for_selector('table', timeout=10000)
        table = page.locator('table')
        expect(table).to_be_visible()
        
        # Get initial table row count  
        initial_rows = table.locator('tr').count()
        assert initial_rows > 0, f"Expected table to have rows, but found {initial_rows} rows"
        
        # Check each imported contact with comprehensive verification
        imported_contacts = [
            {"name": "Test User 1", "email": "test1@example.com", "send_by": "EMAIL"},
            {"name": "Test User 2", "phone": "0504821887", "send_by": "SMS"},
            {"name": "Test User 3", "email": "test3@example.com", "send_by": "EMAIL"}
        ]
        
        contacts_found = 0
        for user_data in imported_contacts:
            # Find contact by name
            contact_name_locator = table.locator(f'text="{user_data["name"]}"')
            try:
                expect(contact_name_locator).to_be_visible(timeout=5000)
                contact_row = contact_name_locator.locator('xpath=ancestor::tr')
                expect(contact_row).to_be_visible()
                
                # Verify contact details based on data
                if 'email' in user_data and user_data['email']:
                    email_cell = contact_row.locator(f'text="{user_data["email"]}"')
                    expect(email_cell).to_be_visible()
                
                if 'send_by' in user_data:
                    send_by_cell = contact_row.locator(f'text="{user_data["send_by"]}"')
                    expect(send_by_cell).to_be_visible()
                
                contacts_found += 1
                allure.attach(f"Successfully verified contact: {user_data['name']}", name=f"Contact Verification", attachment_type=allure.attachment_type.TEXT)
                
            except AssertionError:
                # Take screenshot for debugging this specific contact
                page.screenshot(path=f'artifacts/screenshots/contact_not_found_{user_data["name"].replace(" ", "_")}.png')
                allure.attach(f"Contact not found: {user_data['name']}", name="Missing Contact", attachment_type=allure.attachment_type.TEXT)
                raise AssertionError(f"Contact '{user_data['name']}' was not found in the contacts table after import")
        
        # Final strong assertion - ensure all contacts were found
        assert contacts_found == len(imported_contacts), f"Expected {len(imported_contacts)} contacts to be imported, but only found {contacts_found}"
        allure.attach(f"All {contacts_found} contacts successfully imported and verified", name="Import Success Summary", attachment_type=allure.attachment_type.TEXT)

    @pytest.mark.upload
    @pytest.mark.bilingual
    @allure.story("XLSX Contact Import - Hebrew")
    @allure.title("Import XLSX file with Hebrew names successfully")
    def test_contacts_import_valid_xlsx_file_hebrew_names_success(
        self, bilingual_authenticated_page, test_helpers, test_config, cleanup_test_data
    ):
        """Test importing XLSX file with Hebrew names in Hebrew interface"""
        page, language = bilingual_authenticated_page
        
        if language != 'hebrew':
            pytest.skip("This test only runs in Hebrew interface")
        
        # Navigate to contacts page
        page.goto(f"{test_config.urls['base_url']}dashboard/contacts")
        page.wait_for_load_state('networkidle')
        
        # Create XLSX with Hebrew names
        test_xlsx = self._create_test_xlsx_file([
            {"Full Name": "יוסי כהן", "Email": "yossi@example.com", "Send by": "EMAIL"},
            {"Full Name": "מירי לוי", "Phone": "0552603210", "Send by": "SMS"},
            {"Full Name": "דני אברהם", "Email": "danny@example.com", "Phone": "0551234567", "Send by": "EMAIL"}
        ])
        cleanup_test_data['register_upload'](test_xlsx)
        
        # Import file
        import_button = page.locator('button').filter(has_text='Import Excel')
        import_button.click()
        page.set_input_files('input[type="file"]', test_xlsx)
        page.wait_for_timeout(2000)
        
        # Verify Hebrew names in contacts table
        table = page.locator('table')
        hebrew_names = ["יוסי כהן", "מירי לוי", "דני אברהם"]
        
        for name in hebrew_names:
            contact_row = table.locator(f'text="{name}"')
            expect(contact_row).to_be_visible()
    @pytest.mark.upload
    @pytest.mark.regression
    @allure.story("XLSX Contact Import - Validation")
    @allure.title("Import XLSX file with invalid data shows proper validation")
    def test_contacts_import_invalid_xlsx_validation(
        self, authenticated_page, test_helpers, test_config, cleanup_test_data
    ):
        """Test XLSX import with invalid data shows validation errors"""
        page = authenticated_page
        
        page.goto(f"{test_config.urls['base_url']}dashboard/contacts")
        page.wait_for_load_state('networkidle')
        
        # Create XLSX with invalid data
        invalid_data = test_config.get_invalid_test_data()
        test_xlsx = self._create_test_xlsx_file([
            {"Full Name": "", "Email": "invalid-email", "Phone": "123"},  # Multiple invalid fields
            {"Full Name": "Valid User", "Email": "valid@example.com", "Phone": "0552603210", "Send by": "EMAIL"},  # One valid
            {"Full Name": invalid_data['invalid_names'][0], "Email": invalid_data['invalid_emails'][1]}  # Invalid name and email
        ])
        cleanup_test_data['register_upload'](test_xlsx)
        
        # Import file
        import_button = page.locator('button').filter(has_text='Import Excel')
        import_button.click()
        page.set_input_files('input[type="file"]', test_xlsx)
        page.wait_for_timeout(3000)
        
        # Verify validation error messages
        error_messages = page.locator('.alert-danger, .error-message, .validation-error')
        expect(error_messages).to_be_visible()
        
        # Verify only valid contacts are imported
        table = page.locator('table')
        valid_contact = table.locator('text="Valid User"')
        expect(valid_contact).to_be_visible()

    @pytest.mark.upload
    @pytest.mark.regression
    @allure.story("Contact Seals")
    @allure.title("Add custom seal to contact successfully")
    def test_contacts_add_seal_to_contact_success_english(
        self, authenticated_page, test_helpers, test_config
    ):
        """Test adding custom seal to contact"""
        page = authenticated_page
        
        page.goto(f"{test_config.urls['base_url']}dashboard/contacts")
        page.wait_for_load_state('networkidle')
        
        # First, create a contact if none exists or use existing one
        existing_contact = page.locator('table tbody tr').first
        if existing_contact.count() == 0:
            # Create a new contact first
            add_contact_button = page.locator('text="Add a new contact"')
            add_contact_button.click()
            
            page.fill('textbox[placeholder*="Full Name"]', "Seal Test User")
            page.fill('textbox[placeholder*="Email"]', "sealtest@example.com")
            page.click('button:has-text("Submit")')
            page.wait_for_timeout(2000)
        
        # Click on "Choose File" button for custom seal
        seal_upload_button = page.locator('button:has-text("Choose File")').first
        seal_upload_button.click()
        
        # Upload seal image
        seal_file = test_config.get_file_path('contact_seal')
        if seal_file and os.path.exists(seal_file):
            page.set_input_files('input[type="file"]', seal_file)
            page.wait_for_timeout(2000)
            
            # Verify seal was uploaded (look for success indicator)
            success_indicator = page.locator('.upload-success, .file-uploaded, button:not(:has-text("Choose File"))')
            expect(success_indicator).to_be_visible()

    @pytest.mark.regression
    @allure.story("Contact Tags")
    @allure.title("Add and manage contact tags successfully")
    def test_contacts_add_tags_success_english(
        self, authenticated_page, test_helpers, test_config
    ):
        """Test adding and managing tags for contacts"""
        page = authenticated_page
        
        page.goto(f"{test_config.urls['base_url']}dashboard/contacts")
        page.wait_for_load_state('networkidle')
        
        # Create new contact with tags
        add_contact_button = page.locator('text="Add a new contact"')
        add_contact_button.click()
        
        # Fill contact form
        page.fill('textbox[placeholder*="Full Name"]', "Tagged User")
        page.fill('textbox[placeholder*="Email"]', "tagged@example.com")
        
        # Add search tags
        tags_input = page.locator('textbox').nth(3)  # Search tags field based on form order
        test_tags = "QA, Testing, Automation"
        tags_input.fill(test_tags)
        
        page.click('button:has-text("Submit")')
        page.wait_for_timeout(2000)
        
        # Verify contact appears with tags
        table = page.locator('table')
        contact_row = table.locator('text="Tagged User"').locator('xpath=ancestor::tr')
        expect(contact_row).to_be_visible()
        
        # Check if tags are visible in the contact row
        tags_cell = contact_row.locator('td').nth(4)  # Search tags column
        expect(tags_cell).to_contain_text("QA")

    @pytest.mark.regression
    @allure.story("Contact Validation")
    @allure.title("Validate global phone numbers correctly")
    def test_contacts_add_contact_global_phone_validation(
        self, authenticated_page, test_helpers, test_config
    ):
        """Test validation of international phone numbers"""
        page = authenticated_page
        
        page.goto(f"{test_config.urls['base_url']}dashboard/contacts")
        page.wait_for_load_state('networkidle')
        
        # Test US phone number
        us_phone = test_config.settings.get('united_state_sms_number', '9783475606')
        
        add_contact_button = page.locator('text="Add a new contact"')
        add_contact_button.click()
        
        page.fill('textbox[placeholder*="Full Name"]', "US Contact")
        page.fill('textbox[placeholder*="Phone"]', us_phone)
        
        # Select SMS as send method
        send_by_select = page.locator('combobox').last
        send_by_select.select_option('SMS')
        
        submit_button = page.locator('button:has-text("Submit")')
        expect(submit_button).not_to_be_disabled()
        
        submit_button.click()
        page.wait_for_timeout(2000)
        
        # Verify contact was created successfully
        table = page.locator('table')
        contact_row = table.locator('text="US Contact"')
        expect(contact_row).to_be_visible()

    @pytest.mark.regression
    @allure.story("Contact Management")
    @allure.title("Skip existing contacts during import")
    def test_contacts_import_skip_existing_contacts(
        self, authenticated_page, test_helpers, test_config, cleanup_test_data
    ):
        """Test skipping existing contacts during XLSX import"""
        page = authenticated_page
        
        page.goto(f"{test_config.urls['base_url']}dashboard/contacts")
        page.wait_for_load_state('networkidle')
        
        # First import - create initial contacts
        test_data = [
            {"Full Name": "Existing User", "Email": "existing@example.com", "Send by": "EMAIL"},
            {"Full Name": "Another User", "Email": "another@example.com", "Send by": "EMAIL"}
        ]
        
        initial_xlsx = self._create_test_xlsx_file(test_data)
        cleanup_test_data['register_upload'](initial_xlsx)
        
        import_button = page.locator('button').filter(has_text='Import Excel')
        import_button.click()
        page.set_input_files('input[type="file"]', initial_xlsx)
        page.wait_for_timeout(3000)
        
        # Count initial contacts
        initial_count = page.locator('table tbody tr').count()
        
        # Second import - same data plus new contact
        duplicate_data = test_data + [
            {"Full Name": "New User", "Email": "new@example.com", "Send by": "EMAIL"}
        ]
        
        duplicate_xlsx = self._create_test_xlsx_file(duplicate_data)
        cleanup_test_data['register_upload'](duplicate_xlsx)
        
        import_button.click()
        page.set_input_files('input[type="file"]', duplicate_xlsx)
        page.wait_for_timeout(3000)
        
        # Verify only new contact was added
        final_count = page.locator('table tbody tr').count()
        assert final_count == initial_count + 1, "Should only add new contacts, skip existing ones"
        
        # Verify new contact exists
        new_contact = page.locator('text="New User"')
        expect(new_contact).to_be_visible()

    def _create_test_xlsx_file(self, data: list) -> str:
        """Create temporary XLSX file with test data"""
        # Create temporary file
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        temp_file.close()
        
        # Create workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        # Write headers
        if data:
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                worksheet.cell(row=1, column=col, value=header)
            
            # Write data rows
            for row, item in enumerate(data, 2):
                for col, header in enumerate(headers, 1):
                    worksheet.cell(row=row, column=col, value=item.get(header, ''))
        
        # Save workbook
        workbook.save(temp_file.name)
        workbook.close()
        
        return temp_file.name

    @pytest.mark.regression
    @allure.story("Contact Editing")
    @allure.title("Edit contact information and validate changes")
    def test_contacts_edit_contact_validation_english(
        self, authenticated_page, test_helpers, test_config
    ):
        """Test editing contact information with validation"""
        page = authenticated_page
        
        page.goto(f"{test_config.urls['base_url']}dashboard/contacts")
        page.wait_for_load_state('networkidle')
        
        # Find first editable contact or create one
        edit_button = page.locator('button').filter(has='img').first  # Edit icon button
        
        if edit_button.count() == 0:
            # Create contact first
            add_contact_button = page.locator('text="Add a new contact"')
            add_contact_button.click()
            
            page.fill('textbox[placeholder*="Full Name"]', "Edit Test User")
            page.fill('textbox[placeholder*="Email"]', "edit@example.com")
            page.click('button:has-text("Submit")')
            page.wait_for_timeout(2000)
            
            edit_button = page.locator('button').filter(has='img').first
        
        # Click edit button
        edit_button.click()
        page.wait_for_timeout(1000)
        
        # Modify contact information
        name_field = page.locator('textbox').first
        name_field.clear()
        name_field.fill("Edited User Name")
        
        # Save changes
        save_button = page.locator('button:has-text("Save"), button:has-text("Update"), button:has-text("Submit")')
        save_button.click()
        page.wait_for_timeout(2000)
        
        # Verify changes were saved
        updated_contact = page.locator('text="Edited User Name"')
        expect(updated_contact).to_be_visible()

    @pytest.mark.regression
    @allure.story("Contact Deletion")
    @allure.title("Delete contact successfully")
    def test_contacts_delete_contact_success_english(
        self, authenticated_page, test_helpers, test_config
    ):
        """Test deleting contact"""
        page = authenticated_page
        
        page.goto(f"{test_config.urls['base_url']}dashboard/contacts")
        page.wait_for_load_state('networkidle')
        
        # Create a contact to delete
        add_contact_button = page.locator('text="Add a new contact"')
        add_contact_button.click()
        
        page.fill('textbox[placeholder*="Full Name"]', "Delete Test User")
        page.fill('textbox[placeholder*="Email"]', "delete@example.com")
        page.click('button:has-text("Submit")')
        page.wait_for_timeout(2000)
        
        # Find and click delete button
        contact_row = page.locator('text="Delete Test User"').locator('xpath=ancestor::tr')
        delete_button = contact_row.locator('button').nth(1)  # Second button is usually delete
        delete_button.click()
        
        # Handle confirmation dialog if it appears
        try:
            confirm_button = page.locator('button:has-text("Confirm"), button:has-text("Delete"), button:has-text("Yes")')
            if confirm_button.is_visible():
                confirm_button.click()
        except:
            pass
        
        page.wait_for_timeout(2000)
        
        # Verify contact is deleted
        deleted_contact = page.locator('text="Delete Test User"')
        expect(deleted_contact).not_to_be_visible()

    @pytest.mark.smoke
    @pytest.mark.regression
    @allure.story("Contact Search")
    @allure.title("Search contacts functionality works correctly")
    def test_contacts_search_functionality_english(
        self, authenticated_page, test_helpers, test_config
    ):
        """Test contact search functionality"""
        page = authenticated_page
        
        page.goto(f"{test_config.urls['base_url']}dashboard/contacts")
        page.wait_for_load_state('networkidle')
        
        # Use search box
        search_box = page.locator('searchbox[placeholder*="Search"]')
        
        # Search for existing contact
        search_term = "Aaron"  # Based on contacts seen in the application
        search_box.fill(search_term)
        page.wait_for_timeout(1000)
        
        # Verify search results
        table_rows = page.locator('table tbody tr')
        visible_rows = table_rows.count()
        
        if visible_rows > 0:
            # Check that visible contacts contain search term
            for i in range(min(visible_rows, 3)):  # Check first few results
                row = table_rows.nth(i)
                row_text = row.inner_text()
                assert search_term.lower() in row_text.lower(), f"Search result should contain '{search_term}'"
        
        # Clear search
        search_box.clear()
        page.wait_for_timeout(1000)
        
        # Verify all contacts are shown again
        all_rows = table_rows.count()
        assert all_rows >= visible_rows, "All contacts should be visible when search is cleared"